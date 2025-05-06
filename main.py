import os
import csv
import shutil
import pyodbc
import shelve
import img2pdf
import datetime
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Import SQL credentials
with shelve.open(f'P:/Users/Steven Cox/sql_creds/credentials') as db:
    server = db['server']
    database = db['database']
    username = db['username']
    password = db['password']
print('Credentials created')

# Connect to your database (replace connection string with your actual connection string)
connection_string = f"DRIVER=ODBC Driver 18 for SQL Server; SERVER={server}; DATABASE={database}; ENCRYPT=no;UID={username}; PWD={password}"
db_connection = pyodbc.connect(connection_string)


def query_database(fileno, formatted_comment):
    print(f'Formatted Comment: {formatted_comment}')
    query = None

    # Query for Complaint Package
    if formatted_comment == 'Complaint Pkg':
        query = f"""
                SELECT TOP 1 
                    CASE 
                        WHEN RIGHT(D.DocumentPath, 1) = '\\' THEN D.DocumentPath + D.DocumentName 
                        ELSE D.DocumentPath + '\\' + D.DocumentName 
                    END AS FullPath, 
                    M.FORW_REFNO
                FROM DMSEngine.dbo.IndexForm_CLS I
                JOIN DMSEngine.dbo.Document D ON I.DocumentID = D.DocumentID
                JOIN CLSMI.dbo.MASTER M ON I.FILENO = M.FILENO
                WHERE I.FILENO = ? 
                AND I.LLCode_Field = 'XISSUITR' 
                ORDER BY DDATE_Field DESC
                """

    # Query for Complaint Package-A
    elif formatted_comment == 'Complaint Pkg-A':
        query = f"""
 
                """

    # Query for Garn Bank
    elif formatted_comment == 'Garn Bank':
        query = f"""
                SELECT TOP 1 
                    CASE 
                        WHEN RIGHT(D.DocumentPath, 1) = '\\' THEN D.DocumentPath + D.DocumentName 
                        ELSE D.DocumentPath + '\\' + D.DocumentName 
                    END AS FullPath, 
                    M.FORW_REFNO
                FROM DMSEngine.dbo.IndexForm_CLS I
                JOIN DMSEngine.dbo.Document D ON I.DocumentID = D.DocumentID
                JOIN CLSMI.dbo.MASTER M ON I.FILENO = M.FILENO
                WHERE I.FILENO = ? 
                AND (I.LLCode_Field = 'XIBG' OR I.LLCode_Field = 'XIBG2' OR LOWER(I.CMT) LIKE '%issued bank garn%')
                ORDER BY DDATE_Field DESC
                """

    # Query for Garn POE
    elif formatted_comment == 'Garn POE':
        query = f"""
                SELECT TOP 1 
                    CASE 
                        WHEN RIGHT(D.DocumentPath, 1) = '\\' THEN D.DocumentPath + D.DocumentName 
                        ELSE D.DocumentPath + '\\' + D.DocumentName 
                    END AS FullPath, 
                    M.FORW_REFNO
                FROM DMSEngine.dbo.IndexForm_CLS I
                JOIN DMSEngine.dbo.Document D ON I.DocumentID = D.DocumentID
                JOIN CLSMI.dbo.MASTER M ON I.FILENO = M.FILENO
                WHERE I.FILENO = ? 
                AND (I.LLCode_Field = 'XIWG' OR I.LLCode_Field = 'XIWG2' OR LOWER(I.CMT) LIKE '%issued wage garn%')
                ORDER BY DDATE_Field DESC
                """

    # Query for Garn Tax
    elif formatted_comment == 'Garn Tax':
        query = f"""
                SELECT TOP 1 
                    CASE 
                        WHEN RIGHT(D.DocumentPath, 1) = '\\' THEN D.DocumentPath + D.DocumentName 
                        ELSE D.DocumentPath + '\\' + D.DocumentName 
                    END AS FullPath, 
                    M.FORW_REFNO
                FROM DMSEngine.dbo.IndexForm_CLS I
                JOIN DMSEngine.dbo.Document D ON I.DocumentID = D.DocumentID
                JOIN CLSMI.dbo.MASTER M ON I.FILENO = M.FILENO
                WHERE I.FILENO = ? 
                AND (I.LLCode_Field = 'XITG' OR  I.LLCode_Field = 'XITG2' OR LOWER(I.CMT) LIKE '%issued tax garn%')
                ORDER BY DDATE_Field DESC
                """

    # Query for Affidavit of Service
    elif formatted_comment == 'Affidavit of Service':
        query = f"""
                SELECT TOP 1 
                    CASE 
                        WHEN RIGHT(D.DocumentPath, 1) = '\\' THEN D.DocumentPath + D.DocumentName 
                        ELSE D.DocumentPath + '\\' + D.DocumentName 
                    END AS FullPath, 
                    M.FORW_REFNO
                FROM DMSEngine.dbo.IndexForm_CLS I
                JOIN DMSEngine.dbo.Document D ON I.DocumentID = D.DocumentID
                JOIN CLSMI.dbo.MASTER M ON I.FILENO = M.FILENO
                WHERE I.FILENO = ?
                AND (I.CMT LIKE '%_SERVE_AFFIDAVIT%' OR I.CMT LIKE '%Proof of Service_%' OR I.CMT LIKE '%-Aff-Successful%')
                AND I.CMT NOT LIKE '%NONSERVE_AFFIDAVIT%'
                AND I.CMT NOT LIKE '%Proof of Service to e-File%'
                AND I.CMT NOT LIKE '%SERV BILL%'
                ORDER BY DDATE_Field DESC;
                """

    if query is None:
        raise ValueError("Invalid document type")

    cursor = db_connection.cursor()
    cursor.execute(query, fileno)

    results = cursor.fetchall()

    # Debug output
    print(f"Results for {fileno}: {results}")

    # Ensure results are not empty and have at least 2 columns
    if not results or len(results[0]) < 2:
        print(f"No valid results for fileno {fileno} and comment {formatted_comment}.")
        return None

    return results


def set_file_name(results, formatted_comment):
    if not results or len(results[0]) < 2:
        print("Error: Results are empty or do not contain the expected data.")
        return None

    try:
        forw_refno = results[0][1]
    except IndexError:
        print("Error: Unable to retrieve FORW_REFNO.")
        return None

    current_date = datetime.datetime.now().strftime('%Y%m%d')
    file_name = None

    if formatted_comment == 'Complaint pkg' or formatted_comment == 'Complaint pkg-A':
        file_name = f'PLEADING_{forw_refno}_{current_date}_006691.pdf'
    elif formatted_comment == 'Garn POE':
        file_name = f'GARNPOE_006691_{forw_refno}_01_{current_date}.pdf'
    elif formatted_comment == 'Garn Bank':
        file_name = f'GARNBANK_006691_{forw_refno}_01_{current_date}.pdf'
    elif formatted_comment == 'Garn Tax':
        file_name = f'GARN_006691_{forw_refno}_01_{current_date}.pdf'
    elif formatted_comment == 'Affidavit of Service':
        file_name = f'AFFSERV_{forw_refno}_{current_date}_006691.pdf'

    return file_name


def create_search_cmedi_d656_sheet(output_folder, current_date):
    cmedi_d656_sheet = f"CMEDI_D656_{current_date}.txt"
    cmedi_d656_sheet_path = os.path.join(output_folder, cmedi_d656_sheet)

    if not os.path.exists(cmedi_d656_sheet_path):
        print("Creating CMEDI D613 sheet...")
        with open(cmedi_d656_sheet_path, mode='w', newline='') as file:
            writer = csv.writer(file, delimiter='\t')
            headers = ["595", "H", "FIRM_FILENO", "DCODE", "DCMT", "DQUEUE", "DTIME", "DPRIORITY", "DDATE",
                       "DELETE_FLAG", "DELETE", "#"]
            writer.writerow(headers)

    return cmedi_d656_sheet_path


def parse_date(date_str):
    for fmt in ("%Y-%m-%d %H:%M:%S", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    raise ValueError(f"Date {date_str} is not in a recognized format")


def cmedi_query(fileno, comment, date):
    # Convert the date to a string and strip any leading or trailing whitespace
    date = str(date).strip()

    # Parse the date using the new function
    parsed_date = parse_date(date)

    # Format the date to match the format in the database (YYYY-MM-DD)
    formatted_date = parsed_date.strftime("%Y-%m-%d")

    # Construct the query to delete the diary 613 entry
    query = f"""
            SELECT FILENO, CODE, COMMENT, QUEUE, TIME, PRIORITY, DATE
            FROM CLSMI.dbo.DIARYINT
            WHERE FILENO = '{fileno}'
            AND CODE = '656'
            AND DATE = '{formatted_date}'
            AND COMMENT = '{comment}'
            """
    print(f"Query: {query}")
    cursor = db_connection.cursor()
    cursor.execute(query)
    row = cursor.fetchone()
    print(f"Query Results: {row}")

    return row


def enter_data(cmedi_d656_sheet_path, results):
    # Format the results to include 595 in the first column, D in the second column, Y in the second last column, and # in the last column
    results = list(results)
    results.insert(0, '595')
    results.insert(1, 'D')
    results.append('Y')
    results.append('Y')
    results.append('#')

    print(f"Formatted Results: {results}")

    # Write the results to the next blank row of the CMEDI_D613_{current_date}.txt file
    with open(cmedi_d656_sheet_path, mode='a', newline='') as file:
        writer = csv.writer(file, delimiter='\t')
        writer.writerow(results)
        print("Data entered into the CMEDI D613 sheet")


def main(input_file, output_folder, gui):
    current_date = datetime.datetime.now().strftime('%m-%d-%Y')
    wb = load_workbook(input_file)
    ws = wb.active

    green_fill = PatternFill(start_color="FF98FA98", end_color="FF98FA98", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFE747", end_color="FFFFE747", fill_type="solid")
    red_fill = PatternFill(start_color="FFF54242", end_color="FFF54242", fill_type="solid")
    purple_fill = PatternFill(start_color="FFC4A7E7", end_color="FFC4A7E7", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if not gui.is_running:
            print("Processing stopped as GUI is closed.")
            break

        if row[0].fill.start_color.index in ['FF98FA98', 'FFFFE747', 'FFF54242', 'FFC4A7E7']:
            print("Row already processed")
            continue

        fileno = row[9].value
        comment = row[13].value
        date = row[6].value

        if fileno is None or comment is None:
            continue

        formatted_comments = {
            'Complaint pkg': 'Complaint Pkg',
            'Complaint pkg-a': 'Complaint Pkg-A',
            'Garnishment POE': 'Garn POE',
            'Garnishment Bank': 'Garn Bank',
            'Garn other (tax)': 'Garn Tax',
            'Aff. of service': 'Affidavit of Service',
        }

        comment = comment.strip()
        formatted_comment = formatted_comments.get(comment)
        print(f"Formatted Comment: {formatted_comment}")

        if formatted_comment is None:
            for cell in row:
                if cell.value is not None:
                    cell.fill = yellow_fill
            wb.save(input_file)
            continue

        results = query_database(fileno, formatted_comment)

        if formatted_comment is None:
            for cell in row:
                if cell.value is not None:
                    cell.fill = yellow_fill
            wb.save(input_file)
            continue

        results = query_database(fileno, formatted_comment)

        if results is None:
            for cell in row:
                if cell.value is not None:
                    cell.fill = red_fill
            wb.save(input_file)
            continue

        file_name = set_file_name(results, formatted_comment)
        file_path = results[0][0]
        output_file = os.path.join(output_folder, file_name)
        shutil.copy(file_path, output_file)

        if file_path.endswith('.tif') or file_path.endswith('.tiff'):
            with open(output_file, 'wb') as pdf_file:
                pdf_file.write(img2pdf.convert(file_path))

        cmedi_d656_sheet_path = create_search_cmedi_d656_sheet(output_folder, current_date)
        results = cmedi_query(fileno, comment, date)

        if results is None:
            for cell in row:
                if cell.value is not None:
                    cell.fill = purple_fill
            wb.save(input_file)
            continue
        else:
            for cell in row:
                if cell.value is not None:
                    cell.fill = green_fill
            wb.save(input_file)
            enter_data(cmedi_d656_sheet_path, results)


if __name__ == '__main__':
    input_file = ''
    output_folder = ''

    # Pass the gui instance to the main function
    main(input_file, output_folder)
